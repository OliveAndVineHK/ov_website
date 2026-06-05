import json, os, hashlib, shutil, subprocess, sys
import openpyxl

ROOT = os.getcwd()
def md5(p):
    return hashlib.md5(open(p, "rb").read()).hexdigest()

ext = json.load(open("i18n/proofreading/extracted.json", encoding="utf-8"))
rows = {r["id"]: r for r in ext["rows"]}
files = sorted({r["file"] for r in ext["rows"]})
baseline = {f: md5(f) for f in files}

# pick representative ids
def find(pred): return next(r for r in ext["rows"] if pred(r))
story = find(lambda r: "storyTitle" in r["id"] and r["file"].endswith("pageAboutUtils.ts"))
bullet = find(lambda r: r["kind"] == "bullet")
tern   = find(lambda r: r["kind"] == "ternary")
integ  = find(lambda r: "integrityTranslations.description" in r["id"])

print("test ids:")
for r in (story, bullet, tern, integ):
    print("  ", r["kind"], r["id"], "| ko=", r["ko"][:30], "| en=", r["en"][:30])

# inject sentinels into a test copy of the xlsx
SRC = "i18n/proofreading/Olive-and-Vine_번역검수.xlsx"
TST = "i18n/proofreading/_test.xlsx"
shutil.copy(SRC, TST)
wb = openpyxl.load_workbook(TST)
sent = {
    story["id"]:  ("ko", "검수테스트1,\\n둘째줄"),       # \n must survive
    bullet["id"]: ("ko", "검수테스트2 불렛 항목"),
    tern["id"]:   ("ko", "검수테스트3 화면문구"),
    integ["id"]:  ("en", "ROUNDTRIP <br /> sentinel line."),  # <br/> + EN edit
}
# write into whichever sheet/row holds each id
for name in wb.sheetnames:
    if name.startswith("📋"): continue
    ws = wb[name]
    if (ws.cell(1,1).value or "") != "ID": continue
    for rr in range(2, ws.max_row+1):
        _id = ws.cell(rr,1).value
        if _id in sent:
            lang, val = sent[_id]
            ws.cell(rr, 6 if lang=="en" else 7).value = val
wb.save(TST)

def run(*a):
    r = subprocess.run(a, capture_output=True, text=True)
    if r.returncode != 0:
        print("CMD FAIL", a, r.stdout, r.stderr); sys.exit(1)
    return r.stdout

# xlsx -> edits -> apply
run("python3", "i18n/scripts/xlsx_to_edits.py", TST, "i18n/proofreading/edits.json")
print(run("node", "i18n/scripts/apply.mjs"))

# verify sentinels present
ok = True
checks = {
    story["file"]:  "검수테스트1,\\n둘째줄",
    bullet["file"]: "검수테스트2 불렛 항목",
    tern["file"]:   "검수테스트3 화면문구",
    integ["file"]:  "ROUNDTRIP <br /> sentinel line.",
}
for f, needle in checks.items():
    body = open(f, encoding="utf-8").read()
    present = needle in body
    ok = ok and present
    print(f"  sentinel in {f}: {present}")

# revert using originals
revert = [
    {"id": story["id"],  "newKo": story["ko"]},
    {"id": bullet["id"], "newKo": bullet["ko"]},
    {"id": tern["id"],   "newKo": tern["ko"]},
    {"id": integ["id"],  "newEn": integ["en"]},
]
json.dump(revert, open("i18n/proofreading/edits.json","w",encoding="utf-8"), ensure_ascii=False)
print(run("node", "i18n/scripts/apply.mjs"))

# compare md5 to baseline
restored = all(md5(f) == baseline[f] for f in files)
print("\nSENTINELS LANDED:", ok)
print("BYTE-IDENTICAL AFTER REVERT:", restored)
if not restored:
    for f in files:
        if md5(f) != baseline[f]: print("  DIFF:", f)

# cleanup
os.remove(TST)
os.remove("i18n/proofreading/edits.json")
print("\nRESULT:", "PASS ✅" if (ok and restored) else "FAIL ❌")
