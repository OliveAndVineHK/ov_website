import json, sys, os
import openpyxl

ROOT = os.getcwd()
xlsx = sys.argv[1] if len(sys.argv) > 1 else "i18n/proofreading/Olive-and-Vine_번역검수.xlsx"
out = sys.argv[2] if len(sys.argv) > 2 else "i18n/proofreading/edits.json"

wb = openpyxl.load_workbook(os.path.join(ROOT, xlsx))
edits = []
for name in wb.sheetnames:
    if name.startswith("📋"):
        continue
    ws = wb[name]
    # header check: A1 should be "ID"
    if (ws.cell(1, 1).value or "") != "ID":
        continue
    for r in range(2, ws.max_row + 1):
        _id = ws.cell(r, 1).value
        if not _id:
            continue
        newEn = ws.cell(r, 6).value
        newKo = ws.cell(r, 7).value
        e = {"id": str(_id)}
        has = False
        if newEn is not None and str(newEn).strip() != "":
            e["newEn"] = str(newEn); has = True
        if newKo is not None and str(newKo).strip() != "":
            e["newKo"] = str(newKo); has = True
        if has:
            edits.append(e)

json.dump(edits, open(os.path.join(ROOT, out), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
print(f"edits: {len(edits)} -> {out}")
