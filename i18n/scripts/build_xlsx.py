import json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

ROOT = os.getcwd()
data = json.load(open(os.path.join(ROOT, "i18n/proofreading/extracted.json"), encoding="utf-8"))
rows = data["rows"]
pages = data["pageOrder"]

FONT = "Arial"
kindKo = {"pair": "문구·라벨", "bullet": "불렛·항목", "ternary": "화면 문구", "meta": "메타(SEO)"}

# ---- styles ----
hdr_lock = PatternFill("solid", fgColor="D9D9D9")      # grey: read-only cols
hdr_edit = PatternFill("solid", fgColor="C6E0B4")      # green: editable cols
hdr_memo = PatternFill("solid", fgColor="FFE699")      # yellow: memo
edit_fill = PatternFill("solid", fgColor="EBF3E3")     # light green body for editable
lock_font = Font(name=FONT, color="595959", size=10)
en_font   = Font(name=FONT, color="1F4E79", size=10)   # EN bluish
ko_font   = Font(name=FONT, color="000000", size=10)
hdr_font  = Font(name=FONT, bold=True, size=10, color="000000")
id_font   = Font(name=FONT, color="A6A6A6", size=8)
wrap_top  = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["ID", "위치", "종류", "EN (원문)", "KO (현재 번역)", "수정 EN", "수정 KO", "메모"]
WIDTHS  = [34, 26, 10, 52, 52, 52, 52, 24]
# locked (read-only) columns 1..5 ; editable 6,7,8
EDIT_COLS = {6, 7, 8}

wb = Workbook()
wb.remove(wb.active)

# ---------- 안내 sheet ----------
guide = wb.create_sheet("📋 안내 (먼저 읽기)")
guide.sheet_view.showGridLines = False
guide_lines = [
    ("Olive & Vine 웹사이트 번역 검수 시트", True, 14),
    ("", False, 11),
    ("■ 사용 방법", True, 12),
    ("1) 아래 탭(시트)이 웹사이트 '페이지' 단위로 나뉘어 있습니다. 라이브 사이트를 보면서 해당 페이지 탭을 여세요.", False, 11),
    ("2) 'EN (원문)'·'KO (현재 번역)' 열은 현재 사이트에 들어 있는 내용입니다. (회색 = 수정하지 마세요)", False, 11),
    ("3) 고칠 내용이 있으면 초록색 '수정 EN' / '수정 KO' 칸에만 새 문장을 적어 주세요.", False, 11),
    ("   - 고칠 게 없는 줄은 비워 두면 됩니다. (빈 칸 = 현재 그대로 유지)", False, 11),
    ("4) 의견·질문은 노란색 '메모' 칸에 자유롭게 적어 주세요.", False, 11),
    ("", False, 11),
    ("■ 꼭 지켜 주세요", True, 12),
    ("· 'ID' 열은 절대 수정·삭제하지 마세요. (이 값으로 코드에 다시 반영합니다)", False, 11),
    ("· 다음 기호는 화면에서 줄바꿈/특수 표시이니 그대로 남겨 두세요:", False, 11),
    ("     \\n  →  줄바꿈      <br> / <br />  →  줄바꿈      ${...}  →  자동으로 바뀌는 값(숫자 등)", False, 11),
    ("· 영문 고유명사 'Olive & Vine', 'Rebecca', 'Mi Young', 'Assurance', 'IFRS' 등은 그대로 두세요.", False, 11),
    ("· 행을 추가·삭제하거나 순서를 바꾸지 마세요. (칸 내용만 채워 주세요)", False, 11),
    ("", False, 11),
    ("■ 열 설명", True, 12),
    ("ID = 코드 위치 식별자(건드리지 마세요)  /  위치 = 페이지 내 대략 위치  /  종류 = 문구·불렛·라벨 등", False, 11),
    ("EN·KO = 현재 내용(읽기용)  /  수정 EN·수정 KO = 새 문장 입력  /  메모 = 의견", False, 11),
    ("", False, 11),
    ("작성이 끝나면 이 파일을 그대로 저장해서 보내 주세요. 자동으로 코드에 반영됩니다.", True, 11),
]
for i, (txt, bold, size) in enumerate(guide_lines, start=1):
    c = guide.cell(row=i, column=1, value=txt)
    c.font = Font(name=FONT, bold=bold, size=size,
                  color="2E5016" if (bold and size >= 12) else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
guide.column_dimensions["A"].width = 120

# ---------- per-page sheets ----------
def sheet_name(page):
    # "05 Corporate 기업 서비스" -> "05 Corporate" (<=31, no invalid chars)
    nm = re.sub(r"[\\/?*\[\]:]", " ", page).strip()
    return nm[:31]

for page in pages:
    prows = [r for r in rows if r["page"] == page]
    if not prows:
        continue
    ws = wb.create_sheet(sheet_name(page))
    ws.sheet_view.showGridLines = False
    # header
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=1 + (ci - 1), value=h)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        c.fill = hdr_edit if ci in (6, 7) else (hdr_memo if ci == 8 else hdr_lock)
    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # body
    for ri, r in enumerate(prows, start=2):
        vals = [r["id"], r["location"], kindKo.get(r["kind"], r["kind"]),
                r["en"], r["ko"], "", "", r.get("note", "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = wrap_top
            c.border = border
            if ci == 1:
                c.font = id_font; c.protection = Protection(locked=True)
            elif ci == 4:
                c.font = en_font; c.protection = Protection(locked=True)
            elif ci == 5:
                c.font = ko_font; c.protection = Protection(locked=True)
            elif ci in EDIT_COLS:
                c.font = ko_font; c.protection = Protection(locked=False)
                c.fill = edit_fill
            else:
                c.font = lock_font; c.protection = Protection(locked=True)
    ws.freeze_panes = "C2"            # freeze header + ID/위치 columns
    ws.auto_filter.ref = f"A1:H{len(prows)+1}"
    # No sheet protection — color coding (grey=read-only, green/yellow=edit) guides
    # the proofreader instead, so there is no password prompt.
    ws.protection.sheet = False

out = os.path.join(ROOT, "i18n/proofreading/Olive-and-Vine_번역검수.xlsx")
wb.save(out)
print("saved:", out)
print("sheets:", len(wb.sheetnames), "| rows:", len(rows))
