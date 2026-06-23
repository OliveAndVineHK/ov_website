import json, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# 인사이트 전용 검수 엑셀 빌더. extracted-insights.json(아티클별 page) -> 아티클별 시트 워크북.
# 본문 build_xlsx.py와 동일 컬럼 구조라 xlsx_to_edits.py / apply.mjs 를 그대로 재사용한다.
ROOT = os.getcwd()
data = json.load(open(os.path.join(ROOT, "i18n/proofreading/extracted-insights.json"), encoding="utf-8"))
rows = data["rows"]
pages = data["pageOrder"]  # 이미 아티클 순서(NN slug)

FONT = "Arial"
kindKo = {"pair": "문구·라벨", "bullet": "불렛·항목", "ternary": "화면 문구", "meta": "메타(SEO)"}

hdr_lock = PatternFill("solid", fgColor="D9D9D9")
hdr_edit = PatternFill("solid", fgColor="C6E0B4")
hdr_memo = PatternFill("solid", fgColor="FFE699")
edit_fill = PatternFill("solid", fgColor="EBF3E3")
lock_font = Font(name=FONT, color="595959", size=10)
en_font   = Font(name=FONT, color="1F4E79", size=10)
ko_font   = Font(name=FONT, color="000000", size=10)
hdr_font  = Font(name=FONT, bold=True, size=10, color="000000")
id_font   = Font(name=FONT, color="A6A6A6", size=8)
wrap_top  = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["ID", "위치", "종류", "EN (원문)", "KO (현재 번역)", "수정 EN", "수정 KO", "메모"]
WIDTHS  = [40, 30, 10, 56, 56, 56, 56, 24]
EDIT_COLS = {6, 7, 8}

wb = Workbook()
wb.remove(wb.active)

guide = wb.create_sheet("📋 안내 (먼저 읽기)")
guide.sheet_view.showGridLines = False
guide_lines = [
    ("Olive & Vine 인사이트(아티클) 번역 검수 시트", True, 14),
    ("", False, 11),
    ("■ 이 파일은?", True, 12),
    ("인사이트 아티클 본문의 한국어를 검수하는 시트입니다. 탭(시트) 하나 = 아티클 하나입니다.", False, 11),
    ("AI가 먼저 발행한 한국어를 사람이 다듬는 단계입니다. 채워서 보내 주시면 자동으로 코드에 반영됩니다.", False, 11),
    ("", False, 11),
    ("■ 사용 방법", True, 12),
    ("1) 아래 탭에서 검수할 아티클을 여세요. (탭 이름 = 번호 + 슬러그)", False, 11),
    ("2) 'EN (원문)'·'KO (현재 번역)' 회색 열은 현재 사이트 내용입니다. 수정하지 마세요.", False, 11),
    ("3) 고칠 내용은 초록색 '수정 KO'(필요 시 '수정 EN') 칸에만 적어 주세요. 빈 칸 = 현재 유지.", False, 11),
    ("4) 의견·질문은 노란색 '메모' 칸에 적어 주세요.", False, 11),
    ("", False, 11),
    ("■ 문체·용어 기준 (i18n/RULES.md · GLOSSARY.md v7)", True, 12),
    ("· 본문은 '~합니다' 문어체. 인칭은 '우리'(겸양 '저희'는 문의/CTA·FAQ 질문에 한정).", False, 11),
    ("· 세무 용어는 사이트 기준: 법인세(Profits Tax) · 소득세(Salaries Tax) · 기장(bookkeeping).", False, 11),
    ("· 전문 약어·법령은 한국어(영문) 병기: 세무국(IRD), 회사조례(Companies Ordinance) 등.", False, 11),
    ("· 'insight'=인사이트, 'compliance'=컴플라이언스(문맥상 규정 준수 허용).", False, 11),
    ("", False, 11),
    ("■ 꼭 지켜 주세요", True, 12),
    ("· 'ID' 열은 절대 수정·삭제하지 마세요. (이 값으로 코드에 다시 반영합니다)", False, 11),
    ("· \\n · <br> / <br /> · ${...} 기호는 화면 줄바꿈/자동값이니 그대로 두세요.", False, 11),
    ("· 영문 고유명사 'Olive & Vine', 'IFRS', 'Profits Tax' 등 영문 라벨은 그대로 두세요.", False, 11),
    ("· 행을 추가·삭제하거나 순서를 바꾸지 마세요. (칸 내용만 채워 주세요)", False, 11),
    ("", False, 11),
    ("작성이 끝나면 이 파일을 그대로 저장해서 보내 주세요. 자동으로 코드에 반영됩니다.", True, 11),
]
for i, (txt, bold, size) in enumerate(guide_lines, start=1):
    c = guide.cell(row=i, column=1, value=txt)
    c.font = Font(name=FONT, bold=bold, size=size,
                  color="2E5016" if (bold and size >= 12) else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
guide.column_dimensions["A"].width = 122

def sheet_name(page):
    nm = re.sub(r"[\\/?*\[\]:]", " ", page).strip()
    return nm[:31]

for page in pages:
    prows = [r for r in rows if r["page"] == page]
    if not prows:
        continue
    ws = wb.create_sheet(sheet_name(page))
    ws.sheet_view.showGridLines = False
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        c.fill = hdr_edit if ci in (6, 7) else (hdr_memo if ci == 8 else hdr_lock)
    for ci, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, r in enumerate(prows, start=2):
        vals = [r["id"], r["location"], kindKo.get(r["kind"], r["kind"]),
                r["en"], r["ko"], "", "", r.get("note", "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = wrap_top
            c.border = border
            if ci == 1:
                c.font = id_font; c.protection = Protection(locked=True)
            elif ci == 4:
                c.font = en_font; c.protection = Protection(locked=True)
            elif ci == 5:
                c.font = ko_font; c.protection = Protection(locked=True)
            elif ci in EDIT_COLS:
                c.font = ko_font; c.protection = Protection(locked=False)
                c.fill = edit_fill
            else:
                c.font = lock_font; c.protection = Protection(locked=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:H{len(prows)+1}"
    ws.protection.sheet = False

out = os.path.join(ROOT, "i18n/proofreading/Olive-and-Vine_인사이트검수.xlsx")
wb.save(out)
print("saved:", out)
print("sheets:", len(wb.sheetnames), "(안내 1 + 아티클", len(wb.sheetnames) - 1, ") | rows:", len(rows))
